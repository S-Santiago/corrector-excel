#
# Santiago Fernández
#

import openpyxl, json
from openpyxl.styles import Border, Side
import os

practica = input("Práctica: ")

alumnos = []

notas = {}

def leer_json(ruta_archivo):
    with open(ruta_archivo, 'r') as archivo_json:
        return json.load(archivo_json)

wb = openpyxl.Workbook()
    
sheet = wb.active
sheet.title = "Calificaciones"

sheet['B2'] = 'Alumno'
sheet['C2'] = 'Calificación'

def corregir_excel(ruta: str, alumno: str):
    global wb

    puntuacion = []

    datos_correccion = leer_json(f"{practica}.json")
    archivo_excel = openpyxl.load_workbook(ruta)

    hoja_correcion = wb.create_sheet(title=alumno)

    hoja_correcion["B2"] = "Pestaña"
    hoja_correcion["C2"] = "Celda"
    hoja_correcion["D2"] = "Valor introducido"
    hoja_correcion["E2"] = "Valor necesario"

    error_i = 0

    for pestaña_nombre, correcciones_pestaña in datos_correccion.items():
        hoja_pestaña = archivo_excel[pestaña_nombre]

        for celda, formula in correcciones_pestaña.items():
            celda_alumno = hoja_pestaña[celda]

            if celda_alumno.value == formula:
                puntuacion.append(1)
            else:
                puntuacion.append(0)

                hoja_correcion.cell(row=error_i+3, column=2).value = hoja_pestaña.title
                hoja_correcion.cell(row=error_i+3, column=3).value = celda
                hoja_correcion.cell(row=error_i+3, column=4).value = celda_alumno.value.replace("=", "")
                hoja_correcion.cell(row=error_i+3, column=5).value = formula.replace("=", "")

                error_i += 1
                
    hoja_correcion.column_dimensions['B'].auto_size = True
    hoja_correcion.column_dimensions['C'].auto_size = True
    hoja_correcion.column_dimensions['D'].auto_size = True
    hoja_correcion.column_dimensions['E'].auto_size = True

    return sum(puntuacion)/len(puntuacion)*10

for carpeta_alumno in os.listdir("Prácticas alumnos"):
    ruta_carpeta_alumno = os.path.join("Prácticas alumnos", carpeta_alumno)
    
    if os.path.isdir(ruta_carpeta_alumno):
        for archivo in os.listdir(ruta_carpeta_alumno):
            ruta_archivo = os.path.join(ruta_carpeta_alumno, archivo)
            
            if os.path.isfile(ruta_archivo):
                nombre_alumno = carpeta_alumno.split("_")[0]
                alumnos.append(nombre_alumno)

                notas[nombre_alumno] = corregir_excel(ruta=ruta_archivo, alumno=nombre_alumno)

for i, alumno in enumerate(alumnos):
    sheet.cell(row=i+3, column=2).value = alumno
    sheet.cell(row=i+3, column=3).value = notas[alumno]

sheet.column_dimensions['B'].auto_size = True
sheet.column_dimensions['C'].auto_size = True

wb.save(f'{practica}_calificaciones.xlsx')
print(f'Se ha creado el archivo "{practica}_calificaciones.xlsx" satisfactoriamente.')
input()
